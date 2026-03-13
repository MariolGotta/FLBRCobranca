"""
Script de correção: SRP parcelado + importação de Outpost
=========================================================
Execute NA VPS após o deploy:  python fix_installments.py

O que faz:
  1. SRP Parcelado — lê a coluna NOV/26 (col 75) da aba SRP procurando
     anotações "X/Y":
     - Cria a parcela X (NOV/26) como SRP pago  (foi pulada pelo import)
     - Cria as parcelas X+1 até Y como SRP não pago (dívidas futuras mensais)

  2. Outpost — lê a aba "Outpost" da mesma planilha:
     - Marca has_outpost=True nos jogadores encontrados
     - Para cada mês com valor 'X' ou data: cria dívida de outpost PAGA
     - Meses com '-': sem outpost naquele mês, ignora

Execute UMA VEZ. Rodar mais de uma vez é seguro — verifica duplicatas.
"""

import os, sys
from datetime import datetime, date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

SRP_FILE = os.path.join(BASE_DIR, 'Lista de Membros da Corp - SRP.xlsx')

# Coluna 75 = NOV/26 (última coluna de meses da aba SRP)
ANNOTATION_COL   = 75
ANNOTATION_MONTH = '2026-11'


def _next_month(year_month_str, n=1):
    """Retorna YYYY-MM avançado n meses."""
    y, m = map(int, year_month_str.split('-'))
    total = y * 12 + (m - 1) + n
    return f'{total // 12:04d}-{(total % 12) + 1:02d}'


# ─────────────────────────────────────────────────────────────────────────────
# 1. SRP Parcelado
# ─────────────────────────────────────────────────────────────────────────────
def fix_installments():
    from models import db, Player, Debt, Setting

    wb = _open_workbook()
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    srp_price    = Setting.get('srp_price')
    now          = datetime.utcnow()
    created_paid = created_future = skipped = not_found = 0

    print("\n=== 1. Correção de SRP Parcelado ===\n")

    for row in rows[1:]:
        name = row[1]
        if not name:
            continue
        name = str(name).strip()

        val = row[ANNOTATION_COL] if ANNOTATION_COL < len(row) else None
        if not val:
            continue
        val = str(val).strip()

        if '/' not in val or not val.split('/')[0].isdigit():
            continue

        parts = val.split('/')
        if len(parts) != 2:
            continue
        try:
            current_inst = int(parts[0])
            total_inst   = int(parts[1])
        except ValueError:
            continue

        player = Player.query.filter_by(name=name).first()
        if not player:
            print(f"  AVISO: '{name}' não encontrado no banco.")
            not_found += 1
            continue

        remaining = total_inst - current_inst
        if remaining <= 0:
            print(f"  {name}: {val} — plano encerrado, nada a gerar.")
            continue

        print(f"  {name}: parcela {current_inst}/{total_inst} - criando {remaining} futura(s)...", end='')

        # Parcela atual (NOV/26 = installment X) — foi pulada pelo import
        if not Debt.query.filter_by(player_id=player.id, debt_type='srp',
                                    month=ANNOTATION_MONTH).first():
            db.session.add(Debt(
                player_id   = player.id,
                debt_type   = 'srp',
                amount      = srp_price,
                description = f'SRP - parcela {current_inst}/{total_inst} (parcelamento)',
                month       = ANNOTATION_MONTH,
                paid        = True,
                paid_at     = now,
            ))
            created_paid += 1

        # Parcelas futuras (X+1 até Y)
        for i in range(1, remaining + 1):
            inst_num   = current_inst + i
            inst_month = _next_month(ANNOTATION_MONTH, i)

            if Debt.query.filter_by(player_id=player.id, debt_type='srp',
                                    month=inst_month).first():
                skipped += 1
                continue

            month_date = datetime.strptime(inst_month + '-01', '%Y-%m-%d').date()
            is_past    = month_date < date.today()

            db.session.add(Debt(
                player_id   = player.id,
                debt_type   = 'srp',
                amount      = srp_price,
                description = f'SRP - parcela {inst_num}/{total_inst} (parcelamento)',
                month       = inst_month,
                paid        = is_past,
                paid_at     = now if is_past else None,
            ))
            created_future += 1

        print(" OK")

    db.session.commit()
    print(f"\n  Parcela NOV/26 criada como paga : {created_paid}")
    print(f"  Parcelas futuras criadas         : {created_future}")
    print(f"  Já existentes (puladas)          : {skipped}")
    print(f"  Jogadores não encontrados        : {not_found}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Outpost
# ─────────────────────────────────────────────────────────────────────────────
def import_outpost():
    from models import db, Player, Debt, Setting

    wb  = _open_workbook()
    ws  = wb['Outpost']
    rows = list(ws.iter_rows(values_only=True))

    # Linha 4 (index 3) = cabeçalho: Player, Sistema, Planeta, Nome, <datas>
    header_row = rows[3]
    date_cols  = {}   # col_index -> 'YYYY-MM'
    for i, v in enumerate(header_row):
        if i >= 4 and isinstance(v, datetime):
            date_cols[i] = v.strftime('%Y-%m')

    outpost_price = Setting.get('outpost_price')
    now           = datetime.utcnow()

    players_marked = set()   # evita marcar has_outpost duas vezes
    created = skipped = not_found = 0

    print("\n=== 2. Importação de Outpost ===\n")
    print(f"  Meses encontrados: {len(date_cols)} ({min(date_cols.values())} ate {max(date_cols.values())})\n")

    for row in rows[4:]:
        name = row[0]
        if not name or str(name).strip() in ('Player', ''):
            continue
        name = str(name).strip()

        player = Player.query.filter_by(name=name).first()
        if not player:
            # Tenta match case-insensitive
            from models import Player as P
            player = P.query.filter(P.name.ilike(name)).first()
        if not player:
            print(f"  AVISO: '{name}' não encontrado no banco.")
            not_found += 1
            continue

        # Marca has_outpost na primeira vez que aparece
        if player.id not in players_marked:
            player.has_outpost = True
            players_marked.add(player.id)

        # Importa cada mês
        for col_idx, month_str in date_cols.items():
            cell = row[col_idx] if col_idx < len(row) else None
            if cell is None:
                continue

            # 'X' ou datetime = pago; '-' ou '' = sem outpost aquele mês
            if isinstance(cell, datetime):
                paid_flag = True
            elif str(cell).strip().upper() == 'X':
                paid_flag = True
            else:
                continue   # '-' ou qualquer outra coisa → pula

            # Evita duplicata
            if Debt.query.filter_by(player_id=player.id, debt_type='outpost',
                                    month=month_str).first():
                skipped += 1
                continue

            db.session.add(Debt(
                player_id   = player.id,
                debt_type   = 'outpost',
                amount      = outpost_price,
                description = f'Outpost - {month_str} (importado)',
                month       = month_str,
                paid        = True,
                paid_at     = now,
            ))
            created += 1

    db.session.commit()

    print(f"  Jogadores com outpost marcados   : {len(players_marked)}")
    print(f"  Dívidas de outpost criadas        : {created}")
    print(f"  Já existentes (puladas)           : {skipped}")
    print(f"  Jogadores não encontrados         : {not_found}")
    if not_found:
        print("  (verifique se os nomes na aba Outpost batem com os do banco)")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _open_workbook():
    try:
        import openpyxl
    except ImportError:
        print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(SRP_FILE):
        print(f"ERRO: Planilha não encontrada: {SRP_FILE}")
        print("Coloque 'Lista de Membros da Corp - SRP.xlsx' na pasta do projeto.")
        sys.exit(1)

    return openpyxl.load_workbook(SRP_FILE, data_only=True)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    from app import create_app
    app = create_app()

    with app.app_context():
        fix_installments()
        import_outpost()

    print("\nConcluído!")
