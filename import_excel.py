"""
Script de importação dos arquivos Excel para o banco de dados FLBR Corp.
Execute UMA VEZ após configurar o sistema pela primeira vez.

Uso: python import_excel.py
"""

import os
import sys
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

# Ensure database directory exists
os.makedirs(os.path.join(BASE_DIR, 'database'), exist_ok=True)

from app import create_app
from models import db, Player, Debt, Setting

SRP_FILE = os.path.join(BASE_DIR, 'Lista de Membros da Corp - SRP.xlsx')
DOCTRINE_FILE = os.path.join(BASE_DIR, 'Players_com_nave_de_doutrina_corrigido.xlsx')


def import_all():
    app = create_app()
    with app.app_context():
        print("=== Importação FLBR Corp ===\n")

        if Player.query.count() > 0:
            answer = input("AVISO: Já existem jogadores no banco. Continuar? (s/n): ")
            if answer.lower() != 's':
                print("Importação cancelada.")
                return

        players_created, debts_created = import_srp_file()
        doctrine_updated = import_doctrine_file(players_created)

        print(f"\nOK Jogadores importados: {len(players_created)}")
        print(f"OK Registros de SRP importados: {debts_created}")
        print(f"OK Naves de doutrina atualizadas: {doctrine_updated}")
        print("\nImportacao concluida!")
        print("Acesse http://localhost:5000 e faca login com nome = senha do personagem.")


def import_srp_file():
    """Import players and SRP history from Excel."""
    try:
        import openpyxl
    except ImportError:
        print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(SRP_FILE):
        print(f"AVISO: Arquivo SRP não encontrado: {SRP_FILE}")
        return {}, 0

    print(f"Lendo: {os.path.basename(SRP_FILE)}")
    wb = openpyxl.load_workbook(SRP_FILE, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, 0

    header = rows[0]
    print(f"  Colunas: {len(header)}, Linhas de dados: {len(rows) - 1}")

    # Identify month columns (col index 5 onward, until they look like month headers)
    month_cols = {}  # index -> 'YYYY-MM'
    for i, h in enumerate(header):
        if h and i >= 5:
            parsed = _parse_month_header(str(h))
            if parsed:
                month_cols[i] = parsed

    print(f"  Meses identificados: {len(month_cols)}")

    players_map = {}  # name -> Player
    debts_created = 0

    for row in rows[1:]:
        if not row[1]:  # skip empty rows
            continue

        name = str(row[1]).strip() if row[1] else None
        if not name:
            continue

        category = str(row[2]).strip() if row[2] else 'Piloto'
        occupation = str(row[3]).strip() if row[3] else None
        account_owner = str(row[4]).strip() if row[4] else None

        # Normalize category
        category = _normalize_category(category)
        if occupation:
            occupation = occupation.upper()

        # Determine join_date from first month with any data
        join_date = _infer_join_date(row, month_cols)

        # Check if player already exists
        existing = Player.query.filter_by(name=name).first()
        if existing:
            players_map[name] = existing
            continue

        player = Player(
            name=name,
            category=category,
            occupation=occupation,
            account_owner=account_owner,
            join_date=join_date,
        )
        player.set_password(name)  # initial password = player name

        # Doctrine ships from SRP file (last 3 columns)
        if len(header) >= 82:
            player.doctrine_ship_1 = _cell_val(row, 79)
            player.doctrine_ship_2 = _cell_val(row, 80)
            player.doctrine_ship_3 = _cell_val(row, 81)

        db.session.add(player)
        db.session.flush()  # get player.id

        players_map[name] = player

        # Import SRP payment history
        for col_idx, month_str in month_cols.items():
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx]
            if cell_val is None:
                continue
            val = str(cell_val).strip()

            if val.lower() in ('pagou', 'pago'):
                # Create paid SRP record
                debt = Debt(
                    player_id=player.id,
                    debt_type='srp',
                    amount=Setting.get('srp_price'),
                    description=f'SRP - {month_str} (importado)',
                    month=month_str,
                    paid=True,
                    paid_at=datetime.utcnow(),
                )
                db.session.add(debt)
                debts_created += 1

    db.session.commit()
    print(f"  OK: {len(players_map)} jogadores processados.")
    return players_map, debts_created


def import_doctrine_file(players_map):
    """Update doctrine ships from the second Excel file."""
    try:
        import openpyxl
    except ImportError:
        return 0

    if not os.path.exists(DOCTRINE_FILE):
        print(f"AVISO: Arquivo de doutrina não encontrado: {DOCTRINE_FILE}")
        return 0

    print(f"Lendo: {os.path.basename(DOCTRINE_FILE)}")
    # Must read formulas (data_only=False) because cells use IMPORTRANGE
    # The fallback values in formulas contain the actual data
    wb = openpyxl.load_workbook(DOCTRINE_FILE, data_only=False)
    ws = wb.active

    import re

    def extract_formula_value(cell_value):
        """Extract fallback value from IMPORTRANGE formula: =IFERROR(dummy,"ACTUAL_VALUE")"""
        if cell_value is None:
            return None
        val = str(cell_value)
        m = re.search(r',"([^"]*)"[)\s]*$', val)
        if m:
            result = m.group(1).strip()
            return result if result and result not in ('-', 'NÃO', 'NAO', '') else None
        if not val.startswith('='):
            return val.strip() or None
        return None

    updated = 0
    for row in ws.iter_rows():
        cells = [extract_formula_value(c.value) for c in row]
        if len(cells) < 3:
            continue

        name = cells[1]  # col B = Player Name
        if not name:
            continue

        player = Player.query.filter_by(name=name).first()
        if not player:
            continue

        # Col F (index 5) = Ship 1, G (6) = Ship 2, H (7) = Ship 3
        ship1 = cells[5] if len(cells) > 5 else None
        ship2 = cells[6] if len(cells) > 6 else None
        ship3 = cells[7] if len(cells) > 7 else None

        if any([ship1, ship2, ship3]):
            if ship1:
                player.doctrine_ship_1 = ship1
            if ship2:
                player.doctrine_ship_2 = ship2
            if ship3:
                player.doctrine_ship_3 = ship3
            updated += 1

    db.session.commit()
    print(f"  OK: {updated} naves de doutrina atualizadas.")
    return updated


def _parse_month_header(header_val):
    """Try to parse a month header like 'JAN/21', 'FEV/22', etc. into 'YYYY-MM'."""
    MONTH_MAP = {
        'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
        'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
        'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12',
        'FEB': '02', 'APR': '04', 'MAY': '05', 'AUG': '08',
        'SEP': '09', 'OCT': '10', 'DEC': '12',
    }

    val = str(header_val).strip().upper()
    if '/' in val:
        parts = val.split('/')
        if len(parts) == 2:
            month_abbr = parts[0][:3]
            year_part = parts[1]
            if month_abbr in MONTH_MAP:
                year = int('20' + year_part) if len(year_part) == 2 else int(year_part)
                return f'{year:04d}-{MONTH_MAP[month_abbr]}'
    return None


def _normalize_category(cat):
    """Normalize category string to known values."""
    cat = cat.strip()
    known = ['Novato', 'Clone', 'Piloto', 'Elite', 'Industrial',
             'Ministro', 'CEO', 'Contador', 'Administrador']
    for k in known:
        if cat.lower() == k.lower():
            return k
    # Partial matches
    cat_lower = cat.lower()
    if 'novato' in cat_lower:
        return 'Novato'
    if 'clone' in cat_lower or 'alt' in cat_lower:
        return 'Clone'
    if 'ministro' in cat_lower:
        return 'Ministro'
    if 'elite' in cat_lower:
        return 'Elite'
    if 'industrial' in cat_lower:
        return 'Industrial'
    if 'ceo' in cat_lower:
        return 'CEO'
    if 'contador' in cat_lower:
        return 'Contador'
    if 'admin' in cat_lower:
        return 'Administrador'
    return 'Piloto'


def _infer_join_date(row, month_cols):
    """
    Infer join date from the first month where the player was in the corp.
    '-' = player was NOT in the corp that month (skip it).
    Any other non-empty value = player was in the corp (use this month).
    """
    NOT_IN_CORP_VALUES = {'-', '', 'none'}
    sorted_months = sorted(month_cols.items(), key=lambda x: x[1])
    for col_idx, month_str in sorted_months:
        if col_idx < len(row) and row[col_idx] is not None:
            val = str(row[col_idx]).strip().lower()
            if val and val not in NOT_IN_CORP_VALUES:
                try:
                    return datetime.strptime(month_str + '-01', '%Y-%m-%d').date()
                except Exception:
                    pass
    return date.today()


def _cell_val(row, idx):
    """Safely get a cell value from a row tuple."""
    if idx < len(row) and row[idx]:
        val = str(row[idx]).strip()
        return val if val and val not in ('-', 'NÃO', 'NAO', '') else None
    return None


if __name__ == '__main__':
    import_all()
