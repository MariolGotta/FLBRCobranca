import os
from flask import Blueprint, render_template, current_app
from flask_login import login_required

doctrine_bp = Blueprint('doctrine', __name__, url_prefix='/doctrine')

DOCTRINE_FILE = 'Doutrina FLBR 2025.xlsx'

SLOT_ORDER = ['HIGH', 'MID', 'LOW', 'RIG', 'DRONE', 'CARGO', 'SUBSYSTEM', 'IMPLANT']

DOCTRINES = [
    {'id': 'hitwarp',  'label': 'Hit & Warp', 'sheet': 'hitwarp',  'icon': 'bi-lightning-charge-fill', 'color': 'warning'},
    {'id': 'irondome', 'label': 'Iron Dome',  'sheet': 'irondome', 'icon': 'bi-shield-fill-check',     'color': 'info'},
    {'id': 'blops',    'label': 'BLOPs',      'sheet': 'BLOPs',    'icon': 'bi-eye-slash-fill',        'color': 'danger'},
]


def _slot_sort_key(slot_name):
    for i, prefix in enumerate(SLOT_ORDER):
        if slot_name.upper().startswith(prefix):
            return (i, slot_name)
    return (99, slot_name)


def _parse_sheet(ws):
    """Parse one doctrine sheet. Returns list of {name, slots}."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    # Row 0: ship names at columns 0, 3, 6, …
    header_row = rows[0]
    ships_order = []
    ship_col_map = {}
    for i, val in enumerate(header_row):
        if val and i % 3 == 0:
            name = str(val).strip()
            if name:
                ships_order.append(name)
                ship_col_map[name] = i

    ships = []
    for ship_name in ships_order:
        col = ship_col_map[ship_name]
        slot_dict = {}

        for row in rows[2:]:  # skip header rows (row 0 = names, row 1 = Slot/item/qnt)
            slot = row[col]     if col     < len(row) else None
            item = row[col + 1] if col + 1 < len(row) else None
            qty  = row[col + 2] if col + 2 < len(row) else None

            if not slot or not item:
                continue
            slot = str(slot).strip()
            item = str(item).strip()
            if not slot or not item or item in ('0', 'None', 'Nada'):
                continue
            try:
                qty_int = int(qty) if qty is not None else 1
            except (ValueError, TypeError):
                qty_int = 1

            slot_dict.setdefault(slot, []).append({'item': item, 'qty': qty_int})

        sorted_slots = sorted(slot_dict.items(), key=lambda x: _slot_sort_key(x[0]))
        ships.append({
            'name': ship_name,
            'slots': [{'slot': s, 'entries': entries} for s, entries in sorted_slots],
        })

    return ships


def _parse_all_doctrines():
    """Returns list of doctrine dicts with their ships, or empty list on error."""
    try:
        import openpyxl
        path = os.path.join(current_app.root_path, DOCTRINE_FILE)
        if not os.path.exists(path):
            return []

        wb = openpyxl.load_workbook(path, data_only=True)
        result = []
        for doc in DOCTRINES:
            if doc['sheet'] not in wb.sheetnames:
                continue
            ships = _parse_sheet(wb[doc['sheet']])
            result.append({**doc, 'ships': ships})
        return result

    except Exception as e:
        current_app.logger.error(f'Error parsing doctrine xlsx: {e}')
        return []


@doctrine_bp.route('/')
@login_required
def view():
    doctrines = _parse_all_doctrines()
    return render_template('doctrine.html', doctrines=doctrines)
